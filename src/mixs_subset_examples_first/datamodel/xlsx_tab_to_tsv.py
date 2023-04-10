import csv
import logging
import click
import click_log
import openpyxl


class ExcelSheetExtractor:
    def __init__(self, log_level: str):
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(log_level)

    def load_workbook(self, xlsx_input: str) -> openpyxl.workbook.workbook.Workbook:
        self.logger.debug(f"Loading workbook from {xlsx_input}")
        return openpyxl.load_workbook(xlsx_input)

    def extract_sheet_to_tsv(self, workbook: openpyxl.workbook.workbook.Workbook, sheet_name: str,
                             tsv_output: str) -> None:
        sheet = workbook[sheet_name]
        self.logger.debug(f"Extracting sheet {sheet_name} to {tsv_output}")

        with open(tsv_output, 'w', newline='') as tsv_file:
            writer = csv.writer(tsv_file, delimiter='\t')

            for row in sheet.rows:
                row_values = [cell.value for cell in row]
                writer.writerow(row_values)


@click.command()
@click.option('--xlsx_input', '-i', required=True, help='Path to xlsx_input XLSX file')
@click.option('--sheet', '-s', required=True, help='Name of sheet to extract')
@click.option('--tsv_output', '-o', required=True, help='Path to tsv_output TSV file')
@click.option('--log_level', '-v', default='INFO', help='Log level')
# , help='Path to tsv_output TSV file')
# @click_log.simple_verbosity_option(default='INFO')
def cli(xlsx_input: str, sheet: str, tsv_output: str, log_level: str) -> None:
    extractor = ExcelSheetExtractor(log_level)
    workbook = extractor.load_workbook(xlsx_input)
    extractor.extract_sheet_to_tsv(workbook, sheet, tsv_output)


if __name__ == '__main__':
    cli()
